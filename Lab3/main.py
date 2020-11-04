from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Color, colors
from bs4 import BeautifulSoup
import requests

# 1 Tworzenie 3 arkuszy oraz zapisanie arkusza jako Michcinski_175IC_B2.xlsx
# wb = Workbook()
# ws = wb.active
# ws.title = "Giełda"
# ws = wb.create_sheet("Linki")
# ws = wb.create_sheet("FilmWeb")
#
# print(wb.sheetnames)
# wb.save(filename="Michcinski_175IC_B2.xlsx")

# 2 Załadowanie pliku, Aktywny arkusz 1: "Giełda"
wb = load_workbook("Michcinski_175IC_B2.xlsx")
wb.active = 0
ws = wb.active

## Lista różnych skrótów firm
lettersList = ["cdr", "pkn", "fte", "hrp", "alr"]
listaDodatkowa = []

for letter in lettersList:
    result = requests.get("https://stooq.pl/q/g/?s="+letter)
    src = result.content
    soup = BeautifulSoup(src, "lxml")
    firma1 = []
    for tr_tag in soup.find_all('tr', {'id': 'f13'}):
        td_tag = tr_tag.find_all("td")
        for i in td_tag:
            if "Kurs" in i.text:
                firma1.append(i.next_sibling.next_element.next_element)
            elif "Zmiana"in i.text:
                firma1.append(i.next_sibling.next_element.text)
            elif "Transakcje" in i.text:
                firma1.append(i.next_sibling.next_element)
                listaDodatkowa.append(firma1)
                firma1=[]

## Nadanie tytułu scrapowanym danym jako listA, lista
listA = [('Nazwa Firmy', 'Kod', "Kurs", "Zmiana", "Zmiana", "Transakcje")]
lista = [("CD PROJEKT RED", "CDR", listaDodatkowa[0][0], listaDodatkowa[0][1], listaDodatkowa[0][2], listaDodatkowa[0][3]),
         ("POLSKI KONCERN NAFTOWY ORLEN SA", "PKN", listaDodatkowa[1][0], listaDodatkowa[1][1], listaDodatkowa[1][2], listaDodatkowa[1][3]),
         ("FABRYKI MEBLI FORTE SA", "FTE", listaDodatkowa[2][0], listaDodatkowa[2][1], listaDodatkowa[2][2], listaDodatkowa[2][3]),
         ("HARPER HYGIENICS SA", "HRP", listaDodatkowa[3][0], listaDodatkowa[3][1], listaDodatkowa[3][2], listaDodatkowa[3][3]),
         ("ALIOR BANK SA", "ALR", listaDodatkowa[4][0], listaDodatkowa[4][1], listaDodatkowa[4][2], listaDodatkowa[4][3])]

for i in lista:
    listA.append(i)
for i in listA:
    ws.append(i)

    ft_h = Font(bold=True)
    a1 = ws['A1']
    a1.font = ft_h
    b1 = ws['B1']
    b1.font = ft_h
    c1 = ws['C1']
    c1.font = ft_h
    d1 = ws['D1']
    d1.font = ft_h
    e1 = ws['E1']
    e1.font = ft_h
    f1 = ws['F1']
    f1.font = ft_h
    wb.save(filename="Michcinski_175IC_B2.xlsx")

# 3 Linki, ponowna zmiana aktywnego arkusza
wb = load_workbook("Michcinski_175IC_B2.xlsx")
wb.active = 1
ws = wb.active

result = requests.get("https://www.google.com/")
src = result.content
soup = BeautifulSoup(src, "lxml")

urls = []
listOfLinks = soup.find_all('a', href=True)
for link in listOfLinks:
    urls.append(link.attrs["href"])

urls = [e for e in urls if e not in ('/preferences?hl=pl', '/advanced_search?hl=pl&authuser=0', '/intl/pl/ads/', '/intl/pl/about.html'
                                     , '/intl/pl/policies/privacy/', '/intl/pl/policies/terms/')]

for i in urls:
    ws.append([i])
    wb.save(filename='Michcinski_175IC_B2.xlsx')

# 4 FilmWeb
wb = load_workbook("Michcinski_175IC_B2.xlsx")
wb.active = 2
ws = wb.active

result = requests.get("https://www.filmweb.pl/film/Joker-2019-810167")
src = result.content
soup = BeautifulSoup(src, "lxml")

rezyser = soup.find('span', {'itemprop': 'name'})
print(rezyser.next_element)

premiera = soup.find('span', {'class': 'block'})
print(premiera.next_element)

boxoffice = soup.find('div', {'class': 'boxoffice'})
print(boxoffice.previous_element, boxoffice.next_element, boxoffice.next_sibling.next_element)

ocena2 = soup.find('span', {'class': 'filmRating__rateValue'})
print(ocena2.next_element)

ft_h = Font(bold=True)
a1 = ws['A1']
a1.font = ft_h
ws['A1'] = "Film"
b1 = ws['B1']
b1.font = ft_h
ws['B1'] = "Joker"


listaFilmWeb = [("Reżyser","Data Premiery", "BoxOffice-Świat","BoxOffice-USA", "BoxOffice-Poza USA", "Ocena Filmu")]
elements = [(rezyser.next_element,premiera.next_element,boxoffice.previous_element, boxoffice.next_element, boxoffice.next_sibling.next_element,ocena2.next_element)]

for i in elements:
    listaFilmWeb.append(i)
for j in listaFilmWeb:
    ws.append(j)
    wb.save(filename='Michcinski_175IC_B2.xlsx')